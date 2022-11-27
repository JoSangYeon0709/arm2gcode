#!/usr/bin/env python3
import sys
import matplotlib.pylab as plt
import math
import numpy as np
import gif
import openpyxl
import copy
import rospy
import moveit_commander
import geometry_msgs.msg
from tf.transformations import *
import time


pi= math.pi

class arm2gcode():
    
    gcode_arr = []
    unit = 10**(-3)
    RC=1
    AC=0
    zp_reset_x = 0
    zp_reset_y = 0
    zp_reset_z = 0
    
    
    joint_state_topic = ['joint_states:=/robot/joint_states']
    moveit_commander.roscpp_initialize(joint_state_topic)
    rospy.init_node('moveit_test', anonymous=True)


    robot = moveit_commander.RobotCommander()
    group = moveit_commander.MoveGroupCommander("left_arm")

    left_current_pose = group.get_current_pose(end_effector_link='left_gripper').pose
    #right_current_pose = group.get_current_pose(end_effector_link='right_gripper').pose
    
    waypoints = []
    waypoints.append(left_current_pose)

    fst_position_x=waypoints[0].position.x
    fst_position_y=waypoints[0].position.y
    fst_position_z=waypoints[0].position.z
    fst_q_rot = quaternion_from_euler (0,pi,0)

    wpose = group.get_current_pose().pose

    wpose.position.x = waypoints[0].position.x
    wpose.position.y = waypoints[0].position.y 
    wpose.position.z = waypoints[0].position.z
    
    wpose.orientation.x = fst_q_rot[0]
    wpose.orientation.y = fst_q_rot[1]
    wpose.orientation.z = fst_q_rot[2]
    wpose.orientation.w = fst_q_rot[3]

    waypoints.append(copy.deepcopy(wpose))

    def __init__(self):

        self.g92(arm2gcode.wpose.position.x, arm2gcode.wpose.position.y, arm2gcode.wpose.position.z)
        self.g17_18_19(0,pi,0)
        self.g90()
        self.g21()

        f = open("/mnt/c/Users/JSY/test_gcode.gcode", 'r', encoding='utf-8-sig')
        gcode_lines = f.readlines()
        f.close()

        for line in gcode_lines:
            line = line.strip()
            if not line == "" :
                gcode_par_val1 = line.split()
                gcode_par_val2 = {}
                for gg in gcode_par_val1:
                    gcode_par_val2[gg[0]] =float(gg[1:])
                arm2gcode.gcode_arr.append(gcode_par_val2)

        self.gcode_classify()
        
    def pprcc(self,r,p_st=None,p_end=None,DoR_s="cw",stp=1,print_info=True,print_graph=False,xl=False):
    
        x_st=p_st[0]
        y_st=p_st[1]
        x_end=p_end[0]
        y_end=p_end[1]
        sing_ctrl = 1
        DoR = DoR_s

        if (DoR_s == "cw" or DoR_s == "CW"):
            DoR_s = -1
        elif (DoR_s == "ccw" or DoR_s == "CCW"):
            DoR_s = 1

        if x_st < x_end :
            p2p_diff_x = 1
        else:
            p2p_diff_x = -1

        if y_st < y_end :    
            p2p_diff_y = 1
        else:
            p2p_diff_y = -1

        DoR_calcul = p2p_diff_x * p2p_diff_y

        if DoR_s != DoR_calcul:
            x_st,y_st = p_end[0],p_end[1]
            x_end,y_end = p_st[0],p_st[1]
            sing_ctrl = -1

        p2p_l = math.sqrt((x_end-x_st)**2 + (y_end-y_st)**2)

        if p2p_l > r*2:
            sys.exit("p2p_l: %f, r*2: %f \nThere is no circle that meets the conditions. \n Or r is minus."%(p2p_l,r*2))
        elif p2p_l == r*2:
            sys.exit("\nIt's half a circle.\nUse I,J,K instead of R in the G code.\nOr enter the midpoint of the semicircle as the endpoint(p_end)")

        theta_1 = abs(math.asin((y_end-y_st)/p2p_l))
        p2pline2cc_l = math.sqrt(r**2-(p2p_l/2)**2)
        theta_2 = abs(math.asin(p2pline2cc_l/r))
        theta_t = theta_1+theta_2

        x_calcul = math.cos(theta_t)*r*p2p_diff_x*sing_ctrl
        y_calcul = math.sin(theta_t)*r*p2p_diff_y*sing_ctrl
        cc_x = x_st + x_calcul
        cc_y = y_st + y_calcul

        stp_r = DoR_s*math.radians(stp)/1000

        if cc_y < p_st[1]:    
            cc2pst_diff_y = 1
        else:
            cc2pst_diff_y = -1

        if cc_y < p_end[1]:
            cc2pend_diff_y = 1
        else:
            cc2pend_diff_y = -1

        pst_ccx=(p_st[0]-cc_x)/r
        pend_ccx=(p_end[0]-cc_x)/r
    
        if pst_ccx > 1:
            pst_ccx =1
        elif pst_ccx < -1:
            pst_ccx =-1
        
        if pend_ccx > 1:
            pend_ccx =1
        elif pend_ccx < -1:
            pend_ccx =-1
        
        pst2cc_angle = math.acos(pst_ccx) * cc2pst_diff_y
        pend2cc_angle = math.acos(pend_ccx) * cc2pend_diff_y

        if DoR_s == -1:      
            if pst2cc_angle < pend2cc_angle :
                pend2cc_angle = (math.radians(360) - pend2cc_angle)*-1
        elif DoR_s == 1:
            if pst2cc_angle > pend2cc_angle :
                pend2cc_angle = math.radians(360) + pend2cc_angle

        c_x=[]
        c_y=[]
        c_xy=[]
        
        for theta in np.arange(pst2cc_angle,pend2cc_angle,stp_r):
            p_xy=[]
            p_x = cc_x + r * math.cos(theta)
            p_y = cc_y + r * math.sin(theta)
            p_z=0

            p_xy.append(p_x)
            p_xy.append(p_y)

            c_x.append(p_x)
            c_y.append(p_y)
            c_xy.append(p_xy)

        x_st=p_st[0]
        y_st=p_st[1]
        x_end=p_end[0]
        y_end=p_end[1]
        c_x_len=len(c_x)

        print('circle calculation completed')

        """-------------------------------------"""
        if print_graph :

            a = math.radians(0)
            b = math.radians(360)
            c = math.radians(0.1)
            zz_x=[]
            zz_y=[]

            for theta in np.arange(a,b,c):
                p_x = cc_x + r * math.cos(theta)
                p_y = cc_y + r * math.sin(theta)
                zz_x.append(p_x)
                zz_y.append(p_y)

            xy_axis=self.draw_c(cc_x,cc_y,r,zz_x,zz_y,x_st,y_st,x_end,y_end,DoR,print_info)
            self.draw_c(cc_x,cc_y,r,c_x,c_y,x_st,y_st,x_end,y_end,DoR,print_info=False,circle_text="",xy_axis=xy_axis)    

            frames = []

            for i in range(0,c_x_len,5000):
                c_xx=c_x[0:i]
                c_yy=c_y[0:i]
                plt.scatter(cc_x,cc_y,s=100,c="black",label="center")
                plt.scatter(x_st,y_st,s=100,c="red",label="p_st")
                plt.scatter(x_end,y_end,s=100,c="green",label="p_end")
                plt.text(cc_x,cc_y-0.08,"("+str(round(cc_x,3))+","+ str(round(cc_y,3))+")", fontdict={'size': 14})
                plt.text(x_st,y_st-0.08,"("+str(round(x_st,3))+","+ str(round(y_st,3))+")", fontdict={'size': 14})
                plt.text(x_end,y_end-0.08,"("+str(round(x_end,3))+","+ str(round(y_end,3))+")", fontdict={'size': 14})
                plt.axis(xy_axis)

                frame = self.make_gif(c_xx,c_yy,cc_x,cc_y,r,xy_axis=xy_axis)
                frames.append(frame)

            gif_dura = len(frames)/300
            gif.save(frames, 'circle.gif', duration=gif_dura)

        if xl == True:
            self.save2xl(cc_x,cc_y,r,c_x,c_y,x_st,y_st,x_end,y_end,DoR,p2p_l,c_x_len)

        if print_graph == False:
            return c_xy

    @gif.frame
    def make_gif(self,c_xx,c_yy,cc_x,cc_y,r,xy_axis=None):
        plt.axis(xy_axis)
        plt.plot(c_xx,c_yy)

    def draw_c(self,cc_x,cc_y, r, c_x, c_y ,x_st, y_st, x_end, y_end,DoR="",print_info=False,circle_text="circle ",xy_axis=None):   
        plt.figure(figsize=(10,10))
        plt.plot(c_x,c_y)
        plt.grid()
        plt.scatter(cc_x,cc_y,s=100,c="black",label=DoR)
        plt.scatter(cc_x,cc_y,s=100,c="black",label="center")
        plt.scatter(x_st,y_st,s=100,c="red",label="p_st")
        plt.scatter(x_end,y_end,s=100,c="green",label="p_end")
        plt.text(cc_x,cc_y-0.08,"("+str(round(cc_x,3))+","+ str(round(cc_y,3))+")", fontdict={'size': 14})
        plt.text(x_st,y_st-0.08,"("+str(round(x_st,3))+","+ str(round(y_st,3))+")", fontdict={'size': 14})
        plt.text(x_end,y_end-0.08,"("+str(round(x_end,3))+","+ str(round(y_end,3))+")", fontdict={'size': 14})
        plt.legend()
        plt.xlabel('x-axis')
        plt.ylabel('y-axis')

        if print_info== True:
            info_text = "\nradius: " + str(r)+"\np_st: ("+str(x_st)+", "+str(y_st)+") "+"\np_end: ("+str(x_end)+", "+str(y_end)+")"+"\ncircle center: ("+str(cc_x)+", "+str(cc_y)+")"+"\ndirection of rotation: "+DoR+"\n\n"
            plt.title(info_text,loc='left', pad=20)
            plt.title('\n'+circle_text+' equation graph')    
        else:
            info_text= '\n '+ circle_text +'equation graph'
            plt.title(info_text)

        if xy_axis :
            plt.axis(xy_axis)
        else:
            return plt.axis('square')
        plt.show()
        plt.close()

    def save2xl(self,cc_x,cc_y, r, c_x, c_y ,x_st, y_st, x_end, y_end,DoR,p2p_l,c_x_len):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(row=1,column=1).value="방향"
        sheet.cell(row=2,column=1).value="점 개수"
        sheet.cell(row=3,column=1).value="시작점 x,y"
        sheet.cell(row=4,column=1).value="끝점 x,y"
        sheet.cell(row=5,column=1).value="원 중심 x,y"
        sheet.cell(row=6,column=1).value="반지름"
        sheet.cell(row=7,column=1).value="점간 거리"

        sheet.cell(row=1,column=2).value=DoR
        sheet.cell(row=2,column=2).value=c_x_len
        sheet.cell(row=3,column=2).value=x_st
        sheet.cell(row=4,column=2).value=x_end
        sheet.cell(row=5,column=2).value=cc_x
        sheet.cell(row=6,column=2).value=r
        sheet.cell(row=7,column=2).value=p2p_l

        sheet.cell(row=3,column=3).value=y_st
        sheet.cell(row=4,column=3).value=y_end
        sheet.cell(row=5,column=3).value=cc_y


        sheet.cell(row=1,column=6).value="circle_x"
        sheet.cell(row=1,column=7).value="circle_y"

        for i in range(0,c_x_len):
            sheet.cell(row=i+2,column=6).value = c_x[i]
            sheet.cell(row=i+2,column=7).value = c_y[i]
        wb.save('circle_xy.xlsx')


    def gcode_classify(self):
        for line in arm2gcode.gcode_arr:
            if line.get('M') == None:
                gcode_fst_str= line['G']
                x=line.get('X')
                y=line.get('Y')
                z=line.get('Z')
                
                if x == None:
                    x= 0
                if y == None:
                    y= 0
                if z == None:
                    z= 0

                if gcode_fst_str == 1:   #
                    self.g01(x,y,z)     

                if gcode_fst_str == 2 or gcode_fst_str == 3  :
                    r = line.get('R')
                    r *= arm2gcode.unit

                    gcode_p_st = []
                    gcode_p_st.append(arm2gcode.wpose.position.x)
                    gcode_p_st.append(arm2gcode.wpose.position.y)

                    gcode_p_end = []
                    gcode_p_end.append((arm2gcode.zp_reset_x*arm2gcode.AC) + (arm2gcode.wpose.position.x * arm2gcode.RC)+(x*arm2gcode.unit))
                    gcode_p_end.append((arm2gcode.zp_reset_y*arm2gcode.AC) + (arm2gcode.wpose.position.y * arm2gcode.RC)+(y*arm2gcode.unit))
                    
                    ''' z
                    gcode_p_st.append(arm2gcode.wpose.position.z)
                    gcode_p_end.append((arm2gcode.zp_reset_z*arm2gcode.AC) + (arm2gcode.wpose.position.z * arm2gcode.RC)+(z*arm2gcode.unit))
                    '''
                    
                    if gcode_fst_str == 2:
                        dor_s="CW"
                    else:
                        dor_s="CCW"
                    self.g02_03(r, gcode_p_st, gcode_p_end, dor_s)

                if gcode_fst_str == 4:

                    if line.get('P')==None:
                        self.g04("X",x)
                    else:
                        p=ine.get('P')
                        self.g04("P",p)

                if gcode_fst_str == 17:   
                    self.g17_18_19(0,pi,0)
                if gcode_fst_str == 18:   
                    self.g17_18_19(0,pi/2,0)
                if gcode_fst_str == 19:   
                    self.g17_18_19(0,pi/2,pi/2)
                if gcode_fst_str == 20:   
                    self.g20()
                if gcode_fst_str == 21:   
                    self.g21()
                if gcode_fst_str == 90:   
                    self.g90()
                if gcode_fst_str == 91:   
                    self.g91()
                if gcode_fst_str == 92:   
                    self.g92(x,y,z)
                if gcode_fst_str == 77:   
                    self.g77()
                    
                x,y,z=0,0,0
    

    def gcode_prvs_crdnt(self): # 
        print('aa')
        
    def gcode_moveit(self,eef_stp=0.01):
        
        (plan, fraction) = arm2gcode.group.compute_cartesian_path(arm2gcode.waypoints, eef_stp, 0.0)
        if not plan.joint_trajectory.points:
            print("[ERROR] No trajectory found")
        else:
            arm2gcode.group.execute(plan)

        arm2gcode.waypoints=[]
    
    def g01(self,x,y,z): # 직선
        arm2gcode.wpose.position.x = (arm2gcode.zp_reset_x*arm2gcode.AC) + (arm2gcode.wpose.position.x * arm2gcode.RC)+(x*arm2gcode.unit)
        arm2gcode.wpose.position.y = (arm2gcode.zp_reset_y*arm2gcode.AC) + (arm2gcode.wpose.position.y * arm2gcode.RC)+(y*arm2gcode.unit)
        arm2gcode.wpose.position.z = (arm2gcode.zp_reset_z*arm2gcode.AC) + (arm2gcode.wpose.position.z * arm2gcode.RC)+(z*arm2gcode.unit)
        arm2gcode.waypoints.append(copy.deepcopy(arm2gcode.wpose))
        self.gcode_moveit()
        
    def g02_03(self, r, gcode_p_st, gcode_p_end, dor_s): # CW or CCW
        
        pprcc_c_xy = self.pprcc(r, gcode_p_st, gcode_p_end, dor_s, stp=500)

        for i in range(0,len(pprcc_c_xy)):
            arm2gcode.wpose.position.x = pprcc_c_xy[i][0]
            arm2gcode.wpose.position.y = pprcc_c_xy[i][1]
            #arm2gcode.wpose.position.z = pprcc_c_xy[i][2]
            arm2gcode.waypoints.append(copy.deepcopy(arm2gcode.wpose))

        arm2gcode.wpose.position.x = gcode_p_end[0]
        arm2gcode.wpose.position.y = gcode_p_end[1]
        #arm2gcode.wpose.position.z = gcode_p_end[2]
        arm2gcode.waypoints.append(copy.deepcopy(arm2gcode.wpose))

        self.gcode_moveit(100)

    def g04(self,sec_selete,num): 
        if sec_selete == 'P':
            time.sleep(num*10**-6)
        elif sec_selete == 'X' or 'U':
            time.sleep(num)
    
    def g17_18_19(self,I,J,K): #  X,Y I,J   
        q_orig = np.array([0,0,0,1])        
        q_rot = quaternion_from_euler ( I, J, K ) 
        #q_new = quaternion_multiply ( q_rot , q_orig )
        arm2gcode.wpose.orientation.x = q_rot[0]
        arm2gcode.wpose.orientation.y = q_rot[1]
        arm2gcode.wpose.orientation.z = q_rot[2]
        arm2gcode.wpose.orientation.w = q_rot[3]
             
    def g20(self): # inch
        arm2gcode.unit= 25.4 * 10**(-3)
        
    def g21(self): # mm
        arm2gcode.unit= 10**(-3)
        
    def g90(self): # 절대
        arm2gcode.RC=0
        arm2gcode.AC=1
        
    def g91(self): # 상대
        arm2gcode.RC=1
        arm2gcode.AC=0
        
    def g92(self,x,y,z): 
        arm2gcode.zp_reset_x = x
        arm2gcode.zp_reset_y = y
        arm2gcode.zp_reset_z = z
    
    
    '''-------------------'''
    def g77(self): #  m
        arm2gcode.unit= 1
    '''-------------------'''

if __name__ == '__main__':
    try:
        arm2gcode()
        print("finish. eef moves to endpoint.\nmaybe...")
    except rospy.ROSInterruptException:
        pass    


